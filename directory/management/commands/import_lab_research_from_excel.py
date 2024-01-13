from django.core.management.base import BaseCommand
from directory.models import Fractions, Unit, LaboratoryMaterial, Researches, ReleationsFT
import json
from openpyxl import load_workbook

from podrazdeleniya.models import Podrazdeleniya
from researches.models import Tubes


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        internal_code, title, material, container, department, laboratory_duration = '', '', '', '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код" in cells:
                    internal_code = cells.index("Код")
                    title = cells.index("Название")
                    material = cells.index("Материал")
                    container = cells.index("Контейнер")
                    department = cells.index("Подразделение")
                    laboratory_duration = cells.index("Готовность")
                    starts = True
            else:
                material_obj = LaboratoryMaterial.objects.filter(title=cells[material]).first()
                department_obj = Podrazdeleniya.objects.filter(title=cells[department]).first()
                research = Researches(
                    title=cells[title],
                    internal_code=cells[internal_code],
                    laboratory_material=material_obj,
                    podrazdeleniye=department_obj,
                    laboratory_duration=cells[laboratory_duration]
                )
                research.save()

                tube = Tubes.objects.filter(title=cells[container]).first()
                relation_f = ReleationsFT(tube=tube)
                relation_f.save()
                fraction = Fractions(
                    research=research,
                    title=cells[title],
                    relation=relation_f

                )
                fraction.save()
                self.stdout.write(f'Услуга добавлена - {research.title}, фракция - {fraction.title}')
